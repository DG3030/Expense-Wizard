
import pandas as pd
import os
import re
from datetime import datetime
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference


def auto_adjust_column_widths(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value is not None:
                    display = f"${cell.value:,.2f}" if isinstance(cell.value, (int, float)) else str(cell.value)
                    max_length = max(max_length, len(display))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2


def format_amount_column(ws, column_letter):
    for row in range(2, ws.max_row + 1):
        cell = ws[f"{column_letter}{row}"]
        if isinstance(cell.value, (int, float)):
            font = cell.font
            cell.number_format = '"$"#,##0.00'
            cell.font = Font(
                name=font.name,
                size=font.size,
                bold=font.bold,
                italic=font.italic,
                vertAlign=font.vertAlign,
                underline=font.underline,
                strike=font.strike,
                color=font.color
            )


def add_pie_chart(ws, start_row, end_row):
    pie = PieChart()
    pie.title = None
    labels = Reference(ws, min_col=1, min_row=start_row, max_row=end_row)
    data = Reference(ws, min_col=2, min_row=start_row, max_row=end_row)
    pie.add_data(data, titles_from_data=False)
    pie.set_categories(labels)
    ws.add_chart(pie, "D2")

    title_row = 18
    ws.merge_cells(f'D{title_row}:K{title_row}')
    cell = ws[f'D{title_row}']
    cell.value = "Expense Breakdown"
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')


def write_period_sheets(df_period, suffix, writer):
    categories = df_period["Category"].dropna().unique()
    summary_data = []
    sheet_queue = []

    payment_total = 0
    expense_total = 0

    for category in categories:
        category_df = df_period[df_period["Category"] == category].copy()
        category_df.sort_values(by="Trans. date", inplace=True)

        total = category_df["Amount"].sum()
        summary_data.append((category, total))

        total_row = pd.DataFrame({
            "Description": ["TOTAL"],
            "Amount": [total]
        })
        final_df = pd.concat([category_df, total_row], ignore_index=True)

        safe_category = re.sub(r'[:\\/*?\[\]]', '-', category[:20])
        sheet_name = f"{safe_category}_{suffix}"[:31]
        sheet_queue.append((sheet_name, final_df))

        if "payment" in category.lower() or "credit" in category.lower():
            payment_total += total
        else:
            expense_total += total

    net = expense_total + payment_total
    summary_df = pd.DataFrame(summary_data, columns=["Expenses", "Total Amount"])
    summary_df.sort_values(by="Total Amount", ascending=False, inplace=True)

    sheet_name = f"Summary_{suffix}"[:31]
    summary_df.to_excel(writer, sheet_name=sheet_name, startrow=6, index=False)
    worksheet = writer.book[sheet_name]

    worksheet.merge_cells('A1:B1')
    worksheet['A1'] = f"Summary for {suffix}"
    worksheet['A1'].alignment = Alignment(horizontal='center')
    worksheet['A1'].font = Font(bold=True)

    for idx, (label, value) in enumerate(zip(["Credit Card Payments", "Expense Total", "Difference"],
                                             [payment_total, expense_total, net]), start=2):
        worksheet[f'A{idx}'] = label
        worksheet[f'B{idx}'] = value
        worksheet[f'A{idx}'].font = Font(bold=True)
        worksheet[f'B{idx}'].font = Font(bold=True)

    net_cell = worksheet['B4']
    net_cell.font = Font(bold=True, color="FF0000" if net > 0 else "008000")

    worksheet['A6'].font = Font(bold=True)
    worksheet['B6'].font = Font(bold=True)

    auto_adjust_column_widths(worksheet)
    format_amount_column(worksheet, "B")
    add_pie_chart(worksheet, 7, 6 + len(summary_df))

    for sheet_name, data in sheet_queue:
        data.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.book[sheet_name]
        auto_adjust_column_widths(ws)
        format_amount_column(ws, "D")


def main_processing_function(folder_path, target_year, target_month, output_folder=None):
    if not output_folder:
        output_folder = os.path.join(folder_path, "CleanStatements")
    os.makedirs(output_folder, exist_ok=True)
    output_path = os.path.join(output_folder, f"Clean_Data_{target_year}-{target_month:02d}.xlsx")

    file_list = [f for f in os.listdir(folder_path) if f.startswith("Discover") and f.endswith(".xlsx")]
    if not file_list:
        raise FileNotFoundError("No Discover .xlsx files found in the selected folder.")

    dataframes = []
    for file in file_list:
        df = pd.read_excel(
            os.path.join(folder_path, file),
            skiprows=11,
            names=["Trans. date", "Post date", "Description", "Amount", "Category"]
        )
        dataframes.append(df)

    df = pd.concat(dataframes, ignore_index=True)
    df = df.dropna(subset=["Trans. date"])
    df["Amount"] = pd.to_numeric(df["Amount"], errors="coerce")
    df["Trans. date"] = pd.to_datetime(df["Trans. date"], errors="coerce").dt.date
    df["Post date"] = pd.to_datetime(df["Post date"], errors="coerce").dt.date
    df = df.dropna(subset=["Trans. date"])

    df = df[
        (pd.to_datetime(df["Trans. date"]).dt.year == target_year) &
        (pd.to_datetime(df["Trans. date"]).dt.month == target_month)
    ]

    if df.empty:
        raise ValueError(f"No transactions found for {target_year}-{target_month:02d}.")

    mid_month = datetime(target_year, target_month, 15).date()
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        first_half = df[df["Trans. date"] < mid_month]
        second_half = df[df["Trans. date"] >= mid_month]
        write_period_sheets(first_half, "First_Half", writer)
        write_period_sheets(second_half, "Second_Half", writer)

    return output_path
