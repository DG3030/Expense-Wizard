import pandas as pd
import os
import re
from datetime import datetime
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.chart import (PieChart, BarChart, DoughnutChart, RadarChart, Reference)



def auto_adjust_column_widths(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value is not None:
                    display = f"${cell.value:,.2f}" if isinstance(cell.value, (int, float)) else str(cell.value)
                    max_length = max(max_length, len(display))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2


def format_amount_column(ws, column_letter):
    for row in range(2, ws.max_row + 1):
        cell = ws[f"{column_letter}{row}"]
        if isinstance(cell.value, (int, float)):
            font = cell.font
            cell.number_format = '"$"#,##0.00'
            cell.font = Font(
                name=font.name,
                size=font.size,
                bold=font.bold,
                italic=font.italic,
                vertAlign=font.vertAlign,
                underline=font.underline,
                strike=font.strike,
                color=font.color
            )




def add_pie_chart(ws, start_row, end_row):
    chart = PieChart()
    chart.title = "Spending Breakdown"
    data = Reference(ws, min_col=2, min_row=start_row, max_row=end_row)
    categories = Reference(ws, min_col=1, min_row=start_row, max_row=end_row)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(categories)
    ws.add_chart(chart, "D2")

def add_bar_chart(ws, start_row, end_row):
    chart = BarChart()
    chart.title = "Spending Breakdown"
    data = Reference(ws, min_col=2, min_row=start_row, max_row=end_row)
    categories = Reference(ws, min_col=1, min_row=start_row, max_row=end_row)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(categories)
    chart.style = 10
    chart.height = 7
    chart.width = 15
    ws.add_chart(chart, "D2")

def add_column_chart(ws, start_row, end_row):
    chart = BarChart()
    chart.type = "col"
    chart.title = "Spending Breakdown"
    data = Reference(ws, min_col=2, min_row=start_row, max_row=end_row)
    categories = Reference(ws, min_col=1, min_row=start_row, max_row=end_row)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(categories)
    ws.add_chart(chart, "D2")

def add_doughnut_chart(ws, start_row, end_row):
    chart = DoughnutChart()
    chart.title = "Spending Breakdown"
    data = Reference(ws, min_col=2, min_row=start_row, max_row=end_row)
    categories = Reference(ws, min_col=1, min_row=start_row, max_row=end_row)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(categories)
    ws.add_chart(chart, "D2")

def add_radar_chart(ws, start_row, end_row):
    chart = RadarChart()
    chart.title = "Spending Breakdown"
    chart.style = 26
    data = Reference(ws, min_col=2, min_row=start_row, max_row=end_row)
    categories = Reference(ws, min_col=1, min_row=start_row, max_row=end_row)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(categories)
    ws.add_chart(chart, "D2")





def write_period_sheets(df_period, suffix, writer, chart_type):
    from pandas.tseries.offsets import DateOffset, MonthBegin

    categories = df_period["Category"].dropna().unique()
    summary_data = []
    sheet_queue = []

    payment_total = 0
    expense_total = 0

    for category in categories:
        category_df = df_period[df_period["Category"] == category].copy()
        category_df.sort_values(by="Trans. date", inplace=True)

        total = category_df["Amount"].sum()
        summary_data.append((category, total))

        total_row = pd.DataFrame({
            "Description": ["TOTAL"],
            "Amount": [total]
        })
        final_df = pd.concat([category_df, total_row], ignore_index=True)

        safe_category = re.sub(r'[:\\/*?\[\]]', '-', category[:20])
        sheet_name = f"{safe_category}_{suffix}"[:31]
        sheet_queue.append((sheet_name, final_df))

        if "payment" in category.lower() or "credit" in category.lower():
            payment_total += total
        else:
            expense_total += total

    net = expense_total + payment_total
    summary_df = pd.DataFrame(summary_data, columns=["Expenses", "Total Amount"])
    summary_df.sort_values(by="Total Amount", ascending=False, inplace=True)

    sheet_name = f"Summary_{suffix}"[:31]
    summary_df.to_excel(writer, sheet_name=sheet_name, startrow=6, index=False)
    worksheet = writer.book[sheet_name]

    worksheet.merge_cells('A1:B1')
    worksheet['A1'] = f"Summary for {suffix}"
    worksheet['A1'].alignment = Alignment(horizontal='center')
    worksheet['A1'].font = Font(bold=True)

    for idx, (label, value) in enumerate(zip(["Credit Card Payments", "Expense Total", "Difference"],
                                             [payment_total, expense_total, net]), start=2):
        worksheet[f'A{idx}'] = label
        worksheet[f'B{idx}'] = value
        worksheet[f'A{idx}'].font = Font(bold=True)
        worksheet[f'B{idx}'].font = Font(bold=True)

    net_cell = worksheet['B4']
    net_cell.font = Font(bold=True, color="FF0000" if net > 0 else "008000")

    worksheet['A6'].font = Font(bold=True)
    worksheet['B6'].font = Font(bold=True)

    auto_adjust_column_widths(worksheet)
    format_amount_column(worksheet, "B")
    chart_type = chart_type.lower() if chart_type else "pie"


    if not summary_df.empty:    
        if chart_type == "pie":
            add_pie_chart(worksheet, 7, 6 + len(summary_df))
        elif chart_type == "bar":
            add_bar_chart(worksheet, 7, 6 + len(summary_df))
        elif chart_type == "column":
            add_column_chart(worksheet, 7, 6 + len(summary_df))
        elif chart_type == "doughnut":
            add_doughnut_chart(worksheet, 7, 6 + len(summary_df))
        elif chart_type == "radar":
            add_radar_chart(worksheet, 7, 6 + len(summary_df))


    for catergory_sheet_name, data in sheet_queue:
        data.to_excel(writer, sheet_name=catergory_sheet_name, index=False)
        ws = writer.book[catergory_sheet_name]
        auto_adjust_column_widths(ws)
        format_amount_column(ws, "D")




def main_processing_function(folder_path, start_date, end_date, output_folder=None, group_mode="Monthly", use_csv=False, chart_type = "Pie"):
    from pandas.tseries.offsets import MonthEnd
    from datetime import timedelta
    import pandas as pd
    import os

    start_date = pd.to_datetime(start_date)
    end_date = pd.to_datetime(end_date)

    # Output folder logic
    if not output_folder:
        output_folder = os.path.join(folder_path, "CleanStatementsCSV" if use_csv else "CleanStatements")
    os.makedirs(output_folder, exist_ok=True)

    # Base file name
    base_filename = f"Sorted_{start_date.strftime('%Y-%m-%d')}_to_{end_date.strftime('%Y-%m-%d')}_{group_mode.lower()}"
    extension = "csv" if use_csv else "xlsx"
    output_path = os.path.join(output_folder, f"{base_filename}.{extension}")

    # Resolve filename conflicts
    counter = 1
    while os.path.exists(output_path):
        output_path = os.path.join(output_folder, f"{base_filename}_copy{counter}.{extension}")
        counter += 1

    # Load Discover .xlsx files
    file_list = [f for f in os.listdir(folder_path) if f.startswith("Discover") and f.endswith(".xlsx")]
    if not file_list:
        raise FileNotFoundError("No Discover .xlsx files found in the selected folder.")

    dataframes = []
    for file in file_list:
        df = pd.read_excel(
            os.path.join(folder_path, file),
            skiprows=11,
            names=["Trans. date", "Post date", "Description", "Amount", "Category"]
        )
        dataframes.append(df)

    df = pd.concat(dataframes, ignore_index=True)
    df = df.dropna(subset=["Trans. date"])
    df["Amount"] = pd.to_numeric(df["Amount"], errors="coerce")
    df["Trans. date"] = pd.to_datetime(df["Trans. date"], format="%m/%d/%Y", errors="coerce")
    df["Post date"] = pd.to_datetime(df["Post date"], format="%m/%d/%Y", errors="coerce")
    df = df.dropna(subset=["Trans. date"])
    df = df[(df["Trans. date"] >= start_date) & (df["Trans. date"] <= end_date)]

    if df.empty:
        raise ValueError("No transactions found in the selected date range.")

    # Build time periods
    periods = []
    current = start_date
    while current <= end_date:
        if group_mode.lower() == "weekly":
            next_end = current + timedelta(days=6)
        elif group_mode.lower() == "biweekly":
            next_end = current.replace(day=15) if current.day <= 15 else (current + MonthEnd(0)).normalize()
        elif group_mode.lower() == "monthly":
            next_end = (current + MonthEnd(0)).normalize()
        else:
            raise ValueError(f"Unsupported grouping mode: {group_mode}")
        periods.append((current, min(next_end, end_date)))
        current = next_end + timedelta(days=1)

    # --- CSV OUTPUT ---
    if use_csv:
        for i, (start, end) in enumerate(periods, 1):
            period_df = df[(df["Trans. date"] >= start) & (df["Trans. date"] <= end)].copy()
            if not period_df.empty:
                period_df.sort_values(by=["Category", "Trans. date"], inplace=True)

                output_rows = []

                for category in period_df["Category"].dropna().unique():
                    cat_df = period_df[period_df["Category"] == category]
                    output_rows.append([f"Category: {category}", "", "", "", ""])
                    output_rows.append(["Trans. date", "Post date", "Description", "Amount", "Category"])
                    for _, row in cat_df.iterrows():
                        output_rows.append([
                            row["Trans. date"].strftime("%Y-%m-%d"),
                            row["Post date"].strftime("%Y-%m-%d") if pd.notnull(row["Post date"]) else "",
                            row["Description"],
                            f"{row['Amount']:.2f}",
                            row["Category"]
                        ])
                    output_rows.append(["", "", "Subtotal:", f"{cat_df['Amount'].sum():.2f}", ""])
                    output_rows.append([])  # spacer

                output_df = pd.DataFrame(output_rows)

                if group_mode.lower() == "biweekly":
                    suffix = "First_Half" if start.day == 1 else "Second_Half"
                    suffix += f"_{start.strftime('%b_%d')}_to_{end.strftime('%b_%d')}"
                elif group_mode.lower() == "monthly":
                    suffix = f"{start.strftime('%B_%Y')}"
                else:
                    suffix = f"Week_{i}_{start.strftime('%b_%d')}_to_{end.strftime('%b_%d')}"

                filename = os.path.join(output_folder, f"{base_filename}_{suffix}.csv")
                output_df.to_csv(filename, index=False, header=False)

        return output_folder

    # --- EXCEL OUTPUT ---
    else:
        from expense_sorter import write_period_sheets
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for i, (start, end) in enumerate(periods, 1):
                period_df = df[(df["Trans. date"] >= start) & (df["Trans. date"] <= end)]
                if not period_df.empty:
                    if group_mode.lower() == "biweekly":
                        suffix = "First Half" if start.day < 15 else "Second Half"
                        suffix += f" ({start.strftime('%b %d')} – {end.strftime('%b %d')})"
                    elif group_mode.lower() == "monthly":
                        suffix = f"{start.strftime('%B %Y')}"
                    else:
                        suffix = f"Week {i} ({start.strftime('%b %d')} – {end.strftime('%b %d')})"
                    write_period_sheets(period_df, suffix, writer, chart_type)

        return output_path


